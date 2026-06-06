from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum, F
from django.utils import timezone
from django.contrib import messages
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .models import *
from .forms import *
from django.contrib.auth import login, authenticate

# ---------- Role-based permissions ----------
class AdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.role == 'admin'

class ManagerOrAdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.role in ['admin', 'manager']

# ---------- Auth Views ----------
class LoginView(FormView):
    template_name = 'account/login.html'
    form_class = LoginForm
    success_url = reverse_lazy('core:dashboard')

    def form_valid(self, form):
        email = form.cleaned_data.get('username')
        password = form.cleaned_data.get('password')
        user = authenticate(self.request, username=email, password=password)
        if user is not None:
            login(self.request, user)
            return super().form_valid(form)
        form.add_error(None, 'Invalid email or password')
        return self.form_invalid(form)

class SignupView(CreateView):
    model = CustomUser
    form_class = SignupForm
    template_name = 'account/signup.html'
    success_url = reverse_lazy('core:login')

    def form_valid(self, form):
        user = form.save(commit=False)
        user.role = 'waiter'
        user.save()
        StaffProfile.objects.create(user=user)
        messages.success(self.request, 'Account created! Please login.')
        return redirect(self.success_url)

# ---------- Dashboard ----------
class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'dashboard/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        orders_today = Order.objects.filter(created_at__date=today)
        context['total_orders_today'] = orders_today.count()
        context['revenue_today'] = orders_today.aggregate(Sum('net_amount'))['net_amount__sum'] or 0
        context['active_orders'] = Order.objects.filter(status__in=['pending', 'preparing']).count()
        context['low_stock_items'] = Ingredient.objects.filter(current_stock__lte=F('low_stock_threshold')).count()

        from datetime import timedelta
        last_6_months = []
        sales_data = []
        for i in range(6):
            month_date = timezone.now().date() - timedelta(days=30*i)
            month_start = month_date.replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            monthly_total = Order.objects.filter(
                created_at__date__gte=month_start,
                created_at__date__lte=month_end,
                status='completed'
            ).aggregate(Sum('net_amount'))['net_amount__sum'] or 0
            last_6_months.append(month_start.strftime('%b %Y'))
            sales_data.append(float(monthly_total))
        context['months'] = list(reversed(last_6_months))
        context['sales_data'] = list(reversed(sales_data))

        top_items = OrderItem.objects.values('menu_item__name').annotate(total_qty=Sum('quantity')).order_by('-total_qty')[:5]
        context['top_items'] = top_items
        return context

# ---------- Order Management (slug-based) ----------
class OrderListView(LoginRequiredMixin, ListView):
    model = Order
    template_name = 'orders/order_list.html'
    context_object_name = 'orders'
    paginate_by = 10

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.GET.get('search')
        status = self.request.GET.get('status')
        order_type = self.request.GET.get('order_type')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if search:
            qs = qs.filter(Q(slug__icontains=search) | Q(table_number__icontains=search))
        if status:
            qs = qs.filter(status=status)
        if order_type:
            qs = qs.filter(order_type=order_type)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['status_choices'] = Order.STATUS_CHOICES
        ctx['order_type_choices'] = Order.ORDER_TYPE_CHOICES
        return ctx

class OrderCreateView(LoginRequiredMixin, CreateView):
    model = Order
    form_class = OrderForm
    template_name = 'orders/order_form.html'
    success_url = reverse_lazy('core:order_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['menu_items'] = MenuItem.objects.filter(is_available=True)
        return ctx

    def form_valid(self, form):
        response = super().form_valid(form)
        menu_item_ids = self.request.POST.getlist('menu_item[]')
        quantities = self.request.POST.getlist('quantity[]')
        prices = self.request.POST.getlist('price[]')
        instructions = self.request.POST.getlist('special_instructions[]')
        for i, item_id in enumerate(menu_item_ids):
            if item_id:
                menu_item = MenuItem.objects.get(id=item_id)
                OrderItem.objects.create(
                    order=self.object,
                    menu_item=menu_item,
                    quantity=int(quantities[i]),
                    price=Decimal(prices[i]),
                    special_instructions=instructions[i] if i < len(instructions) else ''
                )
        self.object.save()
        messages.success(self.request, 'Order created successfully!')
        return response

class OrderUpdateView(LoginRequiredMixin, UpdateView):
    model = Order
    fields = ['status']
    template_name = 'orders/order_form.html'
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    success_url = reverse_lazy('core:order_list')

class OrderDeleteView(LoginRequiredMixin, DeleteView):
    model = Order
    template_name = 'orders/order_confirm_delete.html'
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    success_url = reverse_lazy('core:order_list')

# ---------- Kitchen Queue ----------
class KitchenQueueView(LoginRequiredMixin, ListView):
    template_name = 'kitchen/kitchen_queue.html'
    context_object_name = 'order_items'

    def get_queryset(self):
        return OrderItem.objects.filter(order__status__in=['pending', 'preparing']).select_related('order', 'menu_item').order_by('order__created_at')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['active_orders_count'] = Order.objects.filter(status__in=['pending', 'preparing']).count()
        return ctx

# ---------- Inventory (slug-based) ----------
class InventoryListView(LoginRequiredMixin, ListView):
    model = Ingredient
    template_name = 'inventory/ingredient_list.html'
    context_object_name = 'ingredients'
    paginate_by = 10

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.GET.get('search')
        low = self.request.GET.get('low_stock')
        if search:
            qs = qs.filter(name__icontains=search)
        if low:
            qs = qs.filter(current_stock__lte=F('low_stock_threshold'))
        return qs

class InventoryCreateView(LoginRequiredMixin, ManagerOrAdminRequiredMixin, CreateView):
    model = Ingredient
    form_class = IngredientForm
    template_name = 'inventory/ingredient_form.html'
    success_url = reverse_lazy('core:inventory_list')

class InventoryUpdateView(LoginRequiredMixin, ManagerOrAdminRequiredMixin, UpdateView):
    model = Ingredient
    form_class = IngredientForm
    template_name = 'inventory/ingredient_form.html'
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    success_url = reverse_lazy('core:inventory_list')

class InventoryDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = Ingredient
    template_name = 'inventory/ingredient_confirm_delete.html'
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    success_url = reverse_lazy('core:inventory_list')

# ---------- Staff Management ----------
class StaffListView(LoginRequiredMixin, ManagerOrAdminRequiredMixin, ListView):
    model = StaffProfile
    template_name = 'staff/staff_list.html'
    context_object_name = 'staff_members'
    paginate_by = 10

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.GET.get('search')
        if search:
            qs = qs.filter(Q(user__full_name__icontains=search) | Q(user__email__icontains=search))
        return qs

class StaffCreateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    model = StaffProfile
    form_class = StaffProfileForm
    template_name = 'staff/staff_form.html'
    success_url = reverse_lazy('core:staff_list')

    def form_valid(self, form):
        email = form.cleaned_data['email']
        full_name = form.cleaned_data['full_name']
        password = form.cleaned_data.get('password', 'tempPass123')
        user = CustomUser.objects.create_user(email=email, password=password, full_name=full_name)
        profile = form.save(commit=False)
        profile.user = user
        profile.save()
        messages.success(self.request, f'Staff {full_name} created.')
        return redirect(self.success_url)

class StaffUpdateView(LoginRequiredMixin, AdminRequiredMixin, UpdateView):
    model = StaffProfile
    form_class = StaffProfileForm
    template_name = 'staff/staff_form.html'
    success_url = reverse_lazy('core:staff_list')

    def get_initial(self):
        initial = super().get_initial()
        initial['email'] = self.object.user.email
        initial['full_name'] = self.object.user.full_name
        return initial

    def form_valid(self, form):
        self.object.user.email = form.cleaned_data['email']
        self.object.user.full_name = form.cleaned_data['full_name']
        if form.cleaned_data.get('password'):
            self.object.user.set_password(form.cleaned_data['password'])
        self.object.user.save()
        return super().form_valid(form)

class StaffDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = StaffProfile
    template_name = 'staff/staff_confirm_delete.html'
    success_url = reverse_lazy('core:staff_list')

    def delete(self, request, *args, **kwargs):
        profile = self.get_object()
        user = profile.user
        response = super().delete(request, *args, **kwargs)
        user.delete()
        messages.success(request, 'Staff member deleted.')
        return response

# ---------- Reports & Excel Export ----------
class ReportsView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/reports.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from datetime import timedelta
        months = []
        revenue = []
        for i in range(6):
            month_date = timezone.now().date() - timedelta(days=30*i)
            month_start = month_date.replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            total = Order.objects.filter(created_at__date__gte=month_start, created_at__date__lte=month_end, status='completed').aggregate(Sum('net_amount'))['net_amount__sum'] or 0
            months.append(month_start.strftime('%b %Y'))
            revenue.append(float(total))
        ctx['months'] = list(reversed(months))
        ctx['revenue_data'] = list(reversed(revenue))
        return ctx

def export_orders_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="orders_export.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers = ['Order ID', 'Table', 'Type', 'Status', 'Items', 'Total', 'Tax', 'Net Amount', 'Created At']
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    orders = Order.objects.prefetch_related('items').order_by('-created_at')
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order.slug[:8])
        ws.cell(row=row, column=2, value=order.table_number or '-')
        ws.cell(row=row, column=3, value=order.get_order_type_display())
        ws.cell(row=row, column=4, value=order.get_status_display())
        ws.cell(row=row, column=5, value=order.items.count())
        ws.cell(row=row, column=6, value=float(order.total_amount))
        ws.cell(row=row, column=7, value=float(order.tax))
        ws.cell(row=row, column=8, value=float(order.net_amount))
        ws.cell(row=row, column=9, value=order.created_at.strftime('%Y-%m-%d %H:%M'))

    for col in [6,7,8]:
        for row in range(2, len(orders)+2):
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_len + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(response)
    return response

# ---------- API Endpoints ----------
def api_notifications(request):
    notifications = Notification.objects.filter(is_read=False).order_by('-created_at')[:5]
    data = [{'message': n.message, 'id': n.id} for n in notifications]
    return JsonResponse({'notifications': data})

def api_mark_notification_read(request, pk):
    if request.method == 'POST':
        Notification.objects.filter(pk=pk).update(is_read=True)
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})

def api_kitchen_queue(request):
    items = OrderItem.objects.filter(order__status__in=['pending', 'preparing']).select_related('order', 'menu_item').values(
        'id', 'order__slug', 'order__table_number', 'order__order_type', 'order__created_at',
        'menu_item__name', 'quantity', 'special_instructions', 'order__status'
    )
    for item in items:
        created = item['order__created_at']
        elapsed = (timezone.now() - created).total_seconds() / 60
        item['elapsed_minutes'] = round(elapsed, 1)
    return JsonResponse(list(items), safe=False)

def api_update_order_status(request, slug):
    if request.method == 'POST':
        order = get_object_or_404(Order, slug=slug)
        new_status = request.POST.get('status')
        if new_status in dict(Order.STATUS_CHOICES):
            order.status = new_status
            order.save()
            return JsonResponse({'success': True})
    return JsonResponse({'success': False})